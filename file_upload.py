from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
from firebase_config import get_storage_bucket
from io import BytesIO
import base64
import json
import threading
from openpyxl import load_workbook 
import time
import csv
import os
import requests
from io import StringIO
from models import db, File, UserProcess, DataFrame, User # Add this import at the top
from datetime import datetime, timezone  # Add timezone to imports at the top

# Create Blueprint
file_processing_bp = Blueprint('file_processing', __name__, url_prefix='/api/')

# Configurations
ALLOWED_EXTENSIONS = {'xls', 'xlsx', 'csv'}
bucket = get_storage_bucket()
CHUNK_SIZE = 15 * 1024 * 1024  # 5MB per chunk

from flask import Blueprint, request, jsonify
import threading
import json
from firebase_config import get_storage_bucket

file_processing_bp = Blueprint('file_processing', __name__, url_prefix='/api')

# Get Firebase Storage bucket
bucket = get_storage_bucket()

from flask import request, jsonify
import threading
import sys
import logging

# Enable debugging logs
logging.basicConfig(level=logging.DEBUG)

@file_processing_bp.route("/process-uploaded-files/", methods=["POST"])
def process_uploaded_files():
    """
    Processes multiple uploaded files: downloads via URLs, saves to database, and triggers metadata extraction.
    """
    try:
        data = request.get_json(force=True, silent=False)
        uploaded_files = data.get("uploadedFiles", {})
        email = data.get("email", "").strip()

        if not uploaded_files or not email:
            print("🚨 ERROR: Missing uploadedFiles or email", flush=True)
            return jsonify({"error": "Missing uploadedFiles or email"}), 400

        # Get user from database
        user = User.query.filter_by(email=email).first()
        if not user:
            print(f"🚨 ERROR: User not found for email: {email}", flush=True)
            return jsonify({"error": "User not found"}), 404

        active_threads = []
        file_records = []

        # Process each file asynchronously
        for file_key, file_info in uploaded_files.items():
            try:
                file_name = file_info.get("fileName")
                file_url = file_info.get("fileUrl")

                if not file_name or not file_url:
                    print(f"⚠️ Skipping invalid file entry: {file_info}", flush=True)
                    continue

                # Check if the file exists in storage
                blob = bucket.blob(f"{email}/uploaded_files/{file_name}")
                if not blob.exists():
                    print(f"🚨 ERROR: File '{file_name}' not found in storage.", flush=True)
                    continue

                # Download file content
                file_content = blob.download_as_bytes()
                file_size = len(file_content)
                file_type = 'Excel' if file_name.endswith(('.xls', '.xlsx')) else 'CSV'

                # Create and save file record in database
                try:
                    current_time = datetime.now(timezone.utc)  # Get current time
                    file_uuid = f"{email}_{file_name}_{current_time.strftime('%Y%m%d%H%M%S')}"  # Create unique ID
                    
                    file_record = File(
                        email=email,
                        file_name=file_name,
                        file_size=file_size,
                        file_type=file_type,
                        file_uuid=file_uuid,
                        user_id=user.id  # Add user ID to the record
                    )
                    # Set upload_time after creation
                    file_record.upload_time = current_time
                    
                    db.session.add(file_record)
                    
                    try:
                        record_dict = {
                            'id': file_record.id,
                            'userId': user.id,  # Just include user ID reference
                            'fileName': file_record.file_name,
                            'fileType': file_record.file_type,
                            'uploadDate': current_time.strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        file_records.append(record_dict)
                    except Exception as dict_error:
                        db.session.rollback()
                        continue

                except Exception as e:
                    print(f"🚨 Error processing file record: {e}", flush=True)
                    continue

                # Start metadata generation in a separate thread
                thread = threading.Thread(
                    target=generate_metadata_async, 
                    args=(file_content, file_name, email),
                    daemon=True
                )
                thread.start()
                active_threads.append(thread)

            except Exception as e:
                print(f"🚨 Error processing file {file_info}: {e}", flush=True)
                continue

        # Commit all file records to database
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"🚨 Database error: {str(e)}", flush=True)
            return jsonify({"error": "Failed to save file records"}), 500

        # Wait for all threads to complete or timeout
        for thread in active_threads:
            thread.join(timeout=30)  # 30 second timeout per thread

        return jsonify({
            "message": "Files downloaded & metadata processing started!",
            "files": file_records,  # Simplified file records
            "userId": user.id,  # Just include user ID at top level
            "email": user.email
        }), 200

    except json.JSONDecodeError as e:
        print(f"🚨 JSON parsing failed: {e}", flush=True)
        return jsonify({"error": "Invalid JSON format"}), 400
    except Exception as e:
        print(f"🚨 ERROR: {e}", flush=True)
        return jsonify({"error": "Failed to process files"}), 500

# Utility Functions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
import threading
from io import BytesIO
import csv
from openpyxl import load_workbook

def extract_metadata(file_content, file_type):
    """
    Optimized metadata extraction with column names and data types for each sheet independently.
    """
    metadata = {}

    def detect_column_type(series):
        """Helper function to detect column type including boolean"""
        try:
            non_null = series.dropna()
            if len(non_null) == 0:
                return 'string'
            
            # Check for boolean
            unique_values = set(non_null.astype(str).str.lower())
            boolean_values = {'true', 'false', '1', '0', 'yes', 'no'}
            if unique_values.issubset(boolean_values):
                return 'boolean'
            
            # Check for numeric
            try:
                # First try to convert to numeric
                numeric_series = pd.to_numeric(non_null)
                
                # Check if all values are integers
                if all(numeric_series.apply(lambda x: float(x).is_integer())):
                    return 'integer'
                else:
                    return 'float'
            except (ValueError, TypeError):
                # If numeric conversion fails, check for date
                try:
                    pd.to_datetime(non_null)
                    return 'date'
                except (ValueError, TypeError):
                    return 'string'
        except:
            return 'string'

    if file_type == 'Excel':
        try:
            excel_file = pd.ExcelFile(BytesIO(file_content))
            sheets_metadata = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                )
                
                # Get column types for this sheet
                column_types = {}
                for column in df.columns:
                    column_types[column] = detect_column_type(df[column])

                sheets_metadata[sheet_name] = {
                    'columns': list(df.columns),
                    'columnTypes': column_types,
                    'rowCount': len(df)
                }

            metadata['sheets'] = sheets_metadata
            metadata['fileType'] = 'Excel'

        except Exception as e:
            print(f"Error extracting Excel metadata: {str(e)}")
            metadata['error'] = f"Error extracting Excel metadata: {str(e)}"

    elif file_type == 'CSV':
        try:
            df = pd.read_csv(BytesIO(file_content))
            
            # Get column types
            column_types = {}
            for column in df.columns:
                column_types[column] = detect_column_type(df[column])

            metadata['columns'] = list(df.columns)
            metadata['columnTypes'] = column_types
            metadata['rowCount'] = len(df)
            metadata['fileType'] = 'CSV'

        except Exception as e:
            print(f"Error extracting CSV metadata: {str(e)}")
            metadata['error'] = f"Error extracting CSV metadata: {str(e)}"

    return metadata

import json
from firebase_config import get_storage_bucket

def save_metadata_to_firebase(metadata, file_name, email):
    """
    Save metadata as a JSON file in Firebase Storage.
    """
    try:
        bucket = get_storage_bucket()
        blob = bucket.blob(f"{email}/metadata/{file_name}.json")
        blob.upload_from_string(json.dumps(metadata), content_type='application/json')
        return True
    except Exception as e:
        print(f"Error saving metadata for {file_name}: {str(e)}")
        return False
import pandas as pd
import json
import traceback
from io import BytesIO
from openpyxl import load_workbook

def generate_metadata_async(file_content, file_name, email):
    """
    Generate metadata asynchronously and upload it to Firebase.
    """
    try:        
        # Determine file type and format
        is_excel = file_name.endswith(('.xls', '.xlsx'))
        file_type = 'Excel' if is_excel else 'CSV'
        file_buffer = None

        try:
            # Ensure the file is wrapped in a BytesIO object
            file_buffer = BytesIO(file_content)

            # Extract metadata
            metadata = extract_metadata(file_content, file_type)

            # Save metadata as JSON in Firebase
            metadata_blob = bucket.blob(f"{email}/metadata/{file_name}.json")
            metadata_blob.upload_from_string(json.dumps(metadata), content_type='application/json')
            print(f"✅ Metadata saved for {file_name}")

        finally:
            if file_buffer:
                file_buffer.close()

        print(f"✅ Metadata completed for {file_name}")

    except Exception as e:
        print(f"🚨 Critical Error in generate_metadata_async: {e}")
        traceback.print_exc()

def fetch_metadata_from_firebase(file_name, email):
    """
    Fetch metadata JSON from Firebase Storage.
    """
    try:
        bucket = get_storage_bucket()
        blob = bucket.blob(f"{email}/metadata/{file_name}.json")
        if blob.exists():
            metadata = json.loads(blob.download_as_string())
            return metadata
        return None
    except Exception as e:
        print(f"Error fetching metadata for {file_name}: {str(e)}")
        return None


def extract_basic_metadata(file_content, file_type):
    """
    Extract essential metadata (minimal processing).
    """
    metadata = {"columns": [], "sheets": {}}
    if file_type == 'Excel':
        xls = pd.ExcelFile(BytesIO(file_content))
        metadata['sheets'] = {
            sheet_name: {
                'column_count': len(pd.read_excel(BytesIO(file_content), sheet_name=sheet_name, nrows=1).columns)
            }
            for sheet_name in xls.sheet_names
        }
    elif file_type == 'CSV':
        df = pd.read_csv(BytesIO(file_content), nrows=1)
        metadata['columns'] = df.columns.tolist()
        metadata['column_count'] = len(df.columns)
    return metadata

def multipart_upload_to_firebase(file_content, file_name, email):
    """
    Perform a multipart upload to Firebase Storage with debug logs.
    """
    blob = bucket.blob(f"{email}/uploaded_files/{file_name}")
    file_size = len(file_content)
    start = 0

    try:
        session = blob.create_resumable_upload_session()
        if not session:
            raise ValueError(f"Failed to create resumable session for {file_name}")

        print(f"Resumable session created for {file_name}: {session}")

        while start < file_size:
            end = min(start + CHUNK_SIZE, file_size)
            chunk = file_content[start:end]
            blob.upload_from_string(chunk, session_id=session, content_type='application/octet-stream')
            print(f"Uploaded chunk: {start}-{end} of {file_size} bytes")
            start = end

        return blob.public_url
    except Exception as e:
        print(f"Error during multipart upload for {file_name}: {str(e)}")
        raise

# Routes
@file_processing_bp.route('/upload_files/', methods=['POST'])
def upload_files():
    """
    Handle file uploads, save metadata as JSON in Firebase, and save file records in database.
    """
    try:
        uploaded_files = request.files.getlist("files")
        email = request.headers.get("X-User-Email") or request.json.get("email")

        if not email:
            return jsonify({"error": "Email is required"}), 400

        if not uploaded_files:
            return jsonify({"error": "No files uploaded"}), 400

        uploaded_urls = []
        file_records = []

        for uploaded_file in uploaded_files:
            filename = secure_filename(uploaded_file.filename)
            file_type = 'Excel' if filename.endswith(('.xls', '.xlsx')) else 'CSV'

            # Read file content
            output_buffer = BytesIO()
            uploaded_file.save(output_buffer)
            file_content = output_buffer.getvalue()
            file_size = len(file_content)

            # Save file to Firebase
            blob = bucket.blob(f"{email}/uploaded_files/{filename}")
            blob.upload_from_string(file_content, content_type='application/octet-stream')
            uploaded_urls.append(blob.public_url)

            # Create and save file record in database
            file_record = File(
                email=email,
                file_name=filename,
                file_size=file_size,
                file_type=file_type
            )
            db.session.add(file_record)
            file_records.append(file_record.to_dict())

            # Extract and save metadata asynchronously
            threading.Thread(
                target=generate_metadata_async, 
                args=(file_content, filename, email)
            ).start()

        # Commit all file records to database
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Database error: {str(e)}")
            return jsonify({"error": "Failed to save file records"}), 500

        return jsonify({
            "success": True,
            "uploaded_urls": uploaded_urls,
            "files": file_records  # Include file records in response
        })

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@file_processing_bp.route('/final_upload/', methods=['POST'])
def final_upload():
    """
    Upload processed files to Firebase using Multipart Upload.
    """
    data = request.json
    file_data = data.get('file_data', {})
    email = data.get('email')

    output_files = []
    errors = []

    for filename, details in file_data.items():
        try:
            file_type = details.get('fileType')
            file_content = base64.b64decode(details.get('content'))
            selected_columns = details.get('selected_columns', {})
            selected_sheets = details.get('selected_sheets', {})

            output_buffer = BytesIO()
            if file_type == 'Excel':
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:
                    for sheet_name, columns in selected_sheets.items():
                        df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name)
                        df[columns].to_excel(writer, sheet_name=sheet_name, index=False)
                output_buffer.seek(0)
            elif file_type == 'CSV':
                df = pd.read_csv(BytesIO(file_content))
                df[selected_columns].to_csv(output_buffer, index=False)
                output_buffer.seek(0)

            # Perform multipart upload
            file_url = multipart_upload_to_firebase(output_buffer.getvalue(), filename, email)
            output_files.append(file_url)

        except Exception as e:
            errors.append(f"Error processing {filename}: {str(e)}")

    if errors:
        return jsonify({"success": False, "errors": errors, "uploaded_files": output_files}), 207
    return jsonify({"success": True, "uploaded_files": output_files})

def get_sheet_data_from_file(email, filename, sheet_name=None, page=1, per_page=100):
    """
    Utility function to fetch sheet data from a file in Firebase Storage.
    """
    try:
        if not email or not filename:
            raise ValueError("Email and filename are required")

        # Calculate skip and limit for pagination
        skip = (page - 1) * per_page

        # Get file from Firebase
        blob = bucket.blob(f"{email}/uploaded_files/{filename}")
        if not blob.exists():
            raise FileNotFoundError(f"File {filename} not found")

        file_content = blob.download_as_bytes()
        file_buffer = BytesIO(file_content)

        try:
            # Handle Excel files
            if filename.endswith(('.xlsx', '.xls')):
                if not sheet_name:
                    raise ValueError("Sheet name is required for Excel files")

                # Read the specific sheet
                df = pd.read_excel(
                    file_buffer,
                    sheet_name=sheet_name,
                    skiprows=skip,
                    nrows=per_page
                )
                
                # Get total row count for pagination
                file_buffer.seek(0)
                total_rows = pd.read_excel(file_buffer, sheet_name=sheet_name).shape[0]

            # Handle CSV files
            else:
                # For CSV files, decode the content to string first
                file_content_str = file_content.decode('utf-8')
                df = pd.read_csv(
                    StringIO(file_content_str),
                    skiprows=skip,
                    nrows=per_page
                )
                
                # Get total row count for pagination
                total_rows = sum(1 for _ in StringIO(file_content_str)) - 1  # Subtract header row

            # Convert DataFrame to dictionary
            data = {
                "columns": list(df.columns),
                "rows": df.to_dict('records'),  # Remove fillna('') to preserve data types
                "pagination": {
                    "total_rows": total_rows,
                    "current_page": page,
                    "per_page": per_page,
                    "total_pages": (total_rows + per_page - 1) // per_page
                }
            }

            return data

        finally:
            file_buffer.close()

    except ValueError as e:
        raise ValueError(str(e))
    except FileNotFoundError as e:
        raise FileNotFoundError(str(e))
    except Exception as e:
        raise Exception(f"Error fetching sheet data: {str(e)}")

@file_processing_bp.route('/get-sheet-data/', methods=['GET'])
def get_sheet_data():
    """
    Endpoint to fetch sheet data for a specific file and sheet name.
    
    Query Parameters:
        filename (str): Name of the file
        sheet_name (str, optional): Name of the sheet (for Excel files)
        page (int, optional): Page number for pagination (default: 1)
        per_page (int, optional): Number of items per page (default: 100)
    
    Headers:
        X-User-Email (str): User's email
    
    Returns:
        JSON object containing:
        {
            'data': {
                'columns': list,
                'rows': list,
                'pagination': {
                    'current_page': int,
                    'per_page': int,
                    'total_pages': int,
                    'total_rows': int
                }
            }
        }
    """
    try:
        # Get parameters from query string
        email = request.headers.get("X-User-Email")
        filename = request.args.get('filename')
        sheet_name = request.args.get('sheet_name')
        
        # Convert page and per_page to integers, with defaults
        try:
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('per_page', 100))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid page or per_page parameter'}), 400

        # Validate required parameters
        if not email:
            return jsonify({'error': 'Email header is required'}), 400
        if not filename:
            return jsonify({'error': 'Filename is required'}), 400

        # Use the utility function
        data = get_sheet_data_from_file(
            email=email,
            filename=filename,
            sheet_name=sheet_name,
            page=page,
            per_page=per_page
        )

        return jsonify({'data': data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def store_dataframe_from_file(email, process_id, table_name, file_id, sheet_name=None, description=""):
    """
    Store a DataFrame from an existing file in a process folder structure and save its metadata.
    
    Args:
        email (str): User's email
        process_id (str): ID of the process
        table_name (str): Name for the table/DataFrame
        file_id (str): ID of the source file
        sheet_name (str, optional): Sheet name for Excel files
        description (str, optional): Description of the table
    
    Returns:
        dict: Status of the operation and metadata information
    """
    try:
        if not all([email, process_id, table_name, file_id]):
            raise ValueError("Email, process ID, table name, and file ID are required")

        # Get process by ID
        process = UserProcess.query.filter_by(id=process_id).first()
        if not process:
            raise ValueError(f"Process with ID '{process_id}' not found")

        # Get file record
        file_record = File.query.filter_by(id=file_id).first()
        if not file_record:
            raise ValueError(f"File with ID '{file_id}' not found")

        # Check if DataFrame already exists in this process
        existing_df = DataFrame.query.filter_by(
            process_id=process_id,
            name=table_name
        ).first()

        if existing_df:
            return None
        
        # Get data from file using existing function
        file_data = get_sheet_data_from_file(
            email=email,
            filename=file_record.file_name,
            sheet_name=sheet_name,
            page=1,
            per_page=1000000
        )

        # Convert to DataFrame
        df = pd.DataFrame(file_data["rows"])

        # Create paths using process ID
        base_path = f"{email}/process/{process_id}"
        df_path = f"{base_path}/dataframes/{table_name}"
        metadata_path = f"{base_path}/metadata/{table_name}.json"
        storage_path = f"{df_path}.csv"

        def detect_column_type(series):
            """Helper function to detect column type including boolean"""
            try:
                non_null = series.dropna()
                if len(non_null) == 0:
                    return 'string'
                
                # Check for boolean
                unique_values = set(non_null.astype(str).str.lower())
                boolean_values = {'true', 'false', '1', '0', 'yes', 'no'}
                if unique_values.issubset(boolean_values):
                    return 'boolean'
                
                # Check for numeric
                try:
                    # First try to convert to numeric
                    numeric_series = pd.to_numeric(non_null)
                    
                    # Check if all values are integers
                    if all(numeric_series.apply(lambda x: float(x).is_integer())):
                        return 'integer'
                    else:
                        return 'float'
                except (ValueError, TypeError):
                    # If numeric conversion fails, check for date
                    try:
                        pd.to_datetime(non_null)
                        return 'date'
                    except (ValueError, TypeError):
                        return 'string'
            except Exception:
                return 'string'

        # Generate detailed metadata
        metadata = {
            "tableName": table_name,
            "description": description,
            "sourceFileId": file_id,
            "sourceFileName": file_record.file_name,
            "sourceSheet": sheet_name if sheet_name else "N/A",
            "processId": process_id,
            "createdAt": (existing_df.created_at if existing_df else datetime.now(timezone.utc)).strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            "updatedAt": datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            "rowCount": len(df),
            "columnCount": len(df.columns),
            "columns": [],
            "originalFileName": file_record.file_name,  # Add original file name to metadata
            "originalSheetName": sheet_name if sheet_name else "N/A"  # Add original sheet name to metadata
        }

        # Add column information
        for column in df.columns:
            column_type = detect_column_type(df[column])
            column_stats = {
                "name": column,
                "type": column_type,
            }
            metadata["columns"].append(column_stats)

        # Save DataFrame as CSV
        df_buffer = BytesIO()
        df.to_csv(df_buffer, index=False)
        df_buffer.seek(0)

        try:
            if existing_df:
                # Update existing record
                existing_df.row_count = len(df)
                existing_df.column_count = len(df.columns)
                existing_df.updated_at = datetime.now(timezone.utc)
                existing_df.storage_path = storage_path
                existing_df.data_metadata = metadata  # Update the metadata
                dataframe_record = existing_df
            else:
                # Create new DataFrame record with is_originally_uploaded=True
                dataframe_record = DataFrame.create_from_pandas(
                    df=df,
                    process_id=process_id,
                    name=table_name,
                    email=email,
                    storage_path=storage_path,
                    user_id=process.user_id,
                    is_originally_uploaded=True,  # Set to True for uploaded files
                    metadata=metadata  # Pass the metadata to create_from_pandas
                )
                dataframe_record.data_metadata = metadata  # Update the metadata
                db.session.add(dataframe_record)

            # Upload DataFrame as CSV (will overwrite if exists)
            df_blob = bucket.blob(storage_path)
            df_blob.upload_from_file(df_buffer, content_type='text/csv')

            # Save metadata (will overwrite if exists)
            metadata_blob = bucket.blob(metadata_path)
            metadata_blob.upload_from_string(
                json.dumps(metadata, indent=2),
                content_type='application/json'
            )

            db.session.commit()

            return {
                "success": True,
                "message": f"DataFrame {'updated' if existing_df else 'created'} successfully",
                "id": dataframe_record.id,
                "name": dataframe_record.name,
                "isUpdate": existing_df is not None
            }

        except Exception as e:
            db.session.rollback()
            # Clean up the uploaded file if database operation fails
            df_blob.delete()
            raise Exception(f"Failed to save DataFrame record: {str(e)}")

    except Exception as e:
        raise Exception(f"Error storing DataFrame: {str(e)}")

@file_processing_bp.route('/store-dataframe-for-process/', methods=['POST'])
def store_process_data():
    """
    Endpoint to store data from an existing file into a process folder.
    """
    try:
        data = request.get_json()
        email = request.headers.get("X-User-Email")
        
        # Validate user
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        process_id = data.get('processId')
        table_name = data.get('tableName')
        file_id = data.get('fileId')  # Changed from fileName
        sheet_name = data.get('sheetName')  # Optional
        description = data.get('description', '')

        # Verify process belongs to user
        process = UserProcess.query.filter_by(id=process_id, user_id=user.id).first()
        if not process:
            return jsonify({"error": "Process not found or access denied"}), 404

        # Get file record and verify ownership
        file_record = File.query.filter_by(id=file_id, user_id=user.id).first()
        if not file_record:
            return jsonify({"error": "File not found or access denied"}), 404

        result = store_dataframe_from_file(
            email=email,
            process_id=process_id,
            table_name=table_name,
            file_id=file_id,  # Changed from filename
            sheet_name=sheet_name,
            description=description
        )

        if result:
            return jsonify(result)
        else:
            return jsonify({
                "error": f'File with TableId {table_name} already exists.',
                "existingProcess": {
                    "id": process.id,
                    "name": process.process_name,
                    "createdAt": process.created_at.isoformat(),
                    "updatedAt": process.updated_at.isoformat()
                }
            }), 409

    except Exception as e:
        print(f"Error in store_process_data endpoint: {str(e)}")
        return jsonify({"error": str(e)}), 500

def get_process_table_data(email, process_id, table_name, page=1, per_page=100):
    """
    Fetch data from a stored table in the process folder.
    
    Args:
        email (str): User's email
        process_id (str): ID of the process
        table_name (str): Name of the table
        page (int, optional): Page number for pagination, defaults to 1
        per_page (int, optional): Items per page, defaults to 100
    
    Returns:
        dict: Dictionary containing table data, metadata, and pagination info
    """
    try:
        if not all([email, process_id, table_name]):
            raise ValueError("Email, process ID, and table name are required")

        # Get process to find the process name for storage path
        process = UserProcess.query.filter_by(id=process_id).first()
        if not process:
            raise ValueError(f"Process with ID '{process_id}' not found")

        # Define paths using process name from database
        df_path = f"{email}/process/{process.id}/dataframes/{table_name}.csv"
        metadata_path = f"{email}/process/{process.id}/metadata/{table_name}.json"

        # Get metadata
        metadata_blob = bucket.blob(metadata_path)
        if not metadata_blob.exists():
            raise FileNotFoundError(f"Metadata for table {table_name} not found")
        
        metadata = json.loads(metadata_blob.download_as_string())

        # Get DataFrame
        df_blob = bucket.blob(df_path)
        if not df_blob.exists():
            raise FileNotFoundError(f"Table {table_name} not found")

        # Calculate pagination
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page

        # Read CSV file
        df_buffer = BytesIO(df_blob.download_as_bytes())
        df = pd.read_csv(df_buffer)
        
        # Get total rows for pagination
        total_rows = len(df)
        
        # Slice the DataFrame for pagination
        df = df.iloc[start_idx:end_idx]

        # Clean data based on column types - preserve data types
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].fillna(0)  # Use 0 for numeric NaN values
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].fillna(pd.Timestamp('1900-01-01'))  # Use default date for date NaN values
            else:
                df[col] = df[col].fillna('')  # Use empty string for string NaN values

        # Prepare response
        response = {
            "metadata": metadata,
            "data": {
                "columns": list(df.columns),
                "rows": df.values.tolist(),  # Convert to simple list of lists
                "pagination": {
                    "total_rows": total_rows,
                    "current_page": page,
                    "per_page": per_page,
                    "total_pages": (total_rows + per_page - 1) // per_page
                }
            },
            "process": {
                "id": process.id,
                "name": process.process_name
            }
        }

        return response

    except ValueError as e:
        raise ValueError(str(e))
    except FileNotFoundError as e:
        raise FileNotFoundError(str(e))
    except Exception as e:
        raise Exception(f"Error fetching table data: {str(e)}")

@file_processing_bp.route('/get-table-data/', methods=['GET'])
def get_table_data():
    """
    Endpoint to fetch data from a table in a process.
    """
    try:
        # Get parameters from query string
        email = request.headers.get("X-User-Email")
        process_id = request.args.get('processId')  # Changed from processName
        table_name = request.args.get('tableName')
        
        try:
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('perPage', 100))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid page or perPage parameter'}), 400

        # Validate required parameters
        if not email:
            return jsonify({'error': 'Email header is required'}), 400
        if not process_id:
            return jsonify({'error': 'Process ID is required'}), 400
        if not table_name:
            return jsonify({'error': 'Table name is required'}), 400

        # Validate user has access to the process
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404

        process = UserProcess.query.filter_by(id=process_id, user_id=user.id).first()
        if not process:
            return jsonify({'error': 'Process not found or access denied'}), 404

        # Get the data
        result = get_process_table_data(
            email=email,
            process_id=process_id,
            table_name=table_name,
            page=page,
            per_page=per_page
        )

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500
