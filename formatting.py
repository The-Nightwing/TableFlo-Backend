from flask import Blueprint, request, jsonify
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, numbers, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from firebase_config import get_storage_bucket
import traceback
import re
import os
import json
from datetime import datetime, timezone
from models import db
from models import (
    User, 
    UserProcess, 
    DataFrame, 
    DataFrameOperation, 
    OperationType,
    FormattingStep,
)
import numpy as np


formatting_bp = Blueprint('formatting', __name__, url_prefix='/api/formatting/')

def convert_to_argb(hex_color):
    """Convert hex color to aRGB format.
    
    Args:
        hex_color (str): Color in hex format (#RRGGBB or RRGGBB)
    Returns:
        str: Color in aRGB format (FF + RRGGBB)
    """
    if not hex_color:
        return 'FF000000'  # Default to black if no color provided
        
    # Remove # if present
    hex_color = hex_color.lstrip('#')
    
    # Add alpha channel if not present
    if len(hex_color) == 6:
        hex_color = 'FF' + hex_color
    
    return hex_color.upper()

def parse_range(range_str):
    """Parse range string into list of cells/columns/rows.
    
    Args:
        range_str (str): The range string to parse. Can be:
            - Column format: 'A', 'A:D', 'A, C, F'
            - Row format: '1', '1:4', '1, 5, 7'
            - Cell format: 'A2', 'A2:D4', 'A2, B4, D6'
    
    Returns:
        tuple: (list of parsed ranges, range_type)
            range_type can be 'columns', 'rows', or 'cells'
        
    Raises:
        ValueError: If the range string is invalid
    """
    if not range_str:
        return [], 'columns'
        
    parts = [p.strip() for p in range_str.split(',')]
    result = []
    
    # Determine range type from the first part
    first_part = parts[0]
    if ':' in first_part:
        start, end = first_part.split(':')
        start = start.strip()
        end = end.strip()
        
        # Check if it's a cell range (A1:D4)
        if re.match(r'^[A-Z]+\d+$', start) and re.match(r'^[A-Z]+\d+$', end):
            range_type = 'cells'
        # Check if it's a column range (A:D)
        elif re.match(r'^[A-Z]+$', start) and re.match(r'^[A-Z]+$', end):
            range_type = 'columns'
        # Check if it's a row range (1:4)
        elif re.match(r'^\d+$', start) and re.match(r'^\d+$', end):
            range_type = 'rows'
        else:
            raise ValueError(f"Invalid range format: {first_part}. Expected format: A:D, 1:4, or A1:D4")
    else:
        # Check if it's a single cell (A1)
        if re.match(r'^[A-Z]+\d+$', first_part):
            range_type = 'cells'
        # Check if it's a single column (A)
        elif re.match(r'^[A-Z]+$', first_part):
            range_type = 'columns'
        # Check if it's a single row (1)
        elif re.match(r'^\d+$', first_part):
            range_type = 'rows'
        else:
            raise ValueError(f"Invalid format: {first_part}. Expected format: A, 1, or A1")
    
    # Parse all parts according to the determined range type
    for part in parts:
        if ':' in part:
            start, end = part.split(':')
            start = start.strip()
            end = end.strip()
            
            if range_type == 'cells':
                if not (re.match(r'^[A-Z]+\d+$', start) and re.match(r'^[A-Z]+\d+$', end)):
                    raise ValueError(f"Invalid cell range format: {part}. Expected format: A1:D4")
                result.append(f"{start}:{end}")
            
            elif range_type == 'columns':
                if not (re.match(r'^[A-Z]+$', start) and re.match(r'^[A-Z]+$', end)):
                    raise ValueError(f"Invalid column range format: {part}. Expected format: A:D")
                start_idx = column_index_from_string(start)
                end_idx = column_index_from_string(end)
                result.extend([get_column_letter(i) for i in range(start_idx, end_idx + 1)])
            
            elif range_type == 'rows':
                if not (re.match(r'^\d+$', start) and re.match(r'^\d+$', end)):
                    raise ValueError(f"Invalid row range format: {part}. Expected format: 1:4")
                result.extend(range(int(start), int(end) + 1))
                
        else:
            if range_type == 'cells':
                if not re.match(r'^[A-Z]+\d+$', part):
                    raise ValueError(f"Invalid cell format: {part}. Expected format: A1")
                result.append(part)
            
            elif range_type == 'columns':
                if not re.match(r'^[A-Z]+$', part):
                    raise ValueError(f"Invalid column format: {part}. Expected format: A")
                result.append(part)
            
            elif range_type == 'rows':
                if not re.match(r'^\d+$', part):
                    raise ValueError(f"Invalid row format: {part}. Expected format: 1")
                result.append(int(part))
    
    return result, range_type

def load_preview_file(bucket, email, file_name, sheet_name=None):
    """
    Load a preview file from Firebase Storage.
    """
    preview_path = f"{email}/previews/{file_name}"
    if sheet_name:
        preview_path += f"/{sheet_name}_preview.csv"
    else:
        preview_path += "/preview.csv"

    blob = bucket.blob(preview_path)
    if not blob.exists():
        raise FileNotFoundError(f"Preview file not found at: {preview_path}")

    content = blob.download_as_text()
    df = pd.read_csv(BytesIO(content.encode('utf-8')), na_values=['NA', ''])
    return df.fillna("-")

@formatting_bp.route('/preview/', methods=['POST'])
def preview_formatting():
    try:
        data = request.json
        email = request.headers.get('X-User-Email')
        if not email:
            return jsonify({"error": "Email is required"}), 400

        file_name = data.get('fileName')
        sheet_name = data.get('sheet')
        formatting_config = data.get('formattingConfig')

        if not file_name or not formatting_config:
            return jsonify({"error": "Missing required parameters"}), 400

        # Load preview file from Firebase
        bucket = get_storage_bucket()
        df = load_preview_file(bucket, email, file_name, sheet_name)

        # Get affected ranges based on location type
        location = formatting_config.get('location', {})
        location_type = location.get('type')
        range_str = location.get('range', '')

        # Convert DataFrame to records for preview
        preview_rows = df.head(10).to_dict('records')
        
        # Add formatting information to preview rows
        format_details = formatting_config.get('format', {})
        format_type = formatting_config.get('type')
        
        # Create a list of affected column indices or cell positions
        affected_positions = []
        ranges, range_type = parse_range(range_str)
        
        for range_item in ranges:
            if location_type == 'Columns':
                if ':' in str(range_item):
                    start, end = str(range_item).split(':')
                    start_idx = column_index_from_string(start) - 1
                    end_idx = column_index_from_string(end)
                    affected_positions.extend(range(start_idx, end_idx))
                else:
                    affected_positions.append(column_index_from_string(str(range_item)) - 1)
        
        # Add formatting metadata to preview rows
        formatted_preview = []
        for row in preview_rows:
            formatted_row = {'values': row, 'formatting': {}}
            
            # Add formatting information based on type
            if format_type == 'Font Colour':
                color = convert_to_argb(format_details.get('fontColor', ''))
                formatted_row['formatting'] = {
                    'type': 'font',
                    'color': f'#{color[2:]}',  # Remove alpha channel for frontend
                    'affected_columns': affected_positions
                }
            
            elif format_type == 'Fill Colour':
                color = convert_to_argb(format_details.get('fillColor', ''))
                formatted_row['formatting'] = {
                    'type': 'fill',
                    'color': f'#{color[2:]}',  # Remove alpha channel for frontend
                    'affected_columns': affected_positions
                }
            
            elif format_type == 'Bold':
                formatted_row['formatting'] = {
                    'type': 'bold',
                    'affected_columns': affected_positions
                }
            
            elif format_type == 'Cell Number Format':
                formatted_row['formatting'] = {
                    'type': 'number',
                    'format': format_details.get('numberFormat'),
                    'affected_columns': affected_positions
                }
            
            elif format_type == 'Column Width':
                formatted_row['formatting'] = {
                    'type': 'width',
                    'width': format_details.get('columnWidth'),
                    'affected_columns': affected_positions
                }
            
            elif format_type == 'Conditional Formatting':
                conditional = format_details.get('conditional', {})
                true_color = convert_to_argb(conditional.get('trueFormat', ''))
                false_color = convert_to_argb(conditional.get('falseFormat', ''))
                formatted_row['formatting'] = {
                    'type': 'conditional',
                    'operator': conditional.get('operator'),
                    'value': conditional.get('value'),
                    'trueColor': f'#{true_color[2:]}',
                    'falseColor': f'#{false_color[2:]}',
                    'affected_columns': affected_positions
                }
            
            formatted_preview.append(formatted_row)

        return jsonify({
            "success": True,
            "data": formatted_preview,
            "preview": {
                "affectedCells": ranges,
                "type": format_type,
                "details": format_details
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

def get_column_type(series):
    """Helper function to get column type, prioritizing 'date' detection."""
    try:
        non_null = series.dropna()
        if len(non_null) == 0:
            return 'string'
        
        dtype_str = str(series.dtype)
        if dtype_str.startswith('bool'):
            return 'boolean'
        
        # Detect dates before numeric (supports YYYYMMDD)
        try:
            str_series = non_null.astype(str)
            import re
            date_patterns = [
                r'^\d{4}-\d{2}-\d{2}$',
                r'^\d{2}/\d{2}/\d{4}$',
                r'^\d{4}/\d{2}/\d{2}$',
                r'^\d{2}-\d{2}-\d{4}$',
                r'^\d{4}\d{2}\d{2}$',
                r'^\d{2}\.\d{2}\.\d{4}$',
                r'^\d{4}\.\d{2}\.\d{2}$',
            ]
            date_matches = sum(1 for val in str_series if any(re.match(p, val) for p in date_patterns))
            if date_matches > len(str_series) * 0.7:
                pd.to_datetime(non_null, errors='raise')
                return 'date'
        except Exception:
            pass
        
        if dtype_str.startswith('int'):
            return 'integer'
        if dtype_str.startswith('float'):
            return 'float'
        
        try:
            pd.to_datetime(non_null, errors='raise')
            return 'date'
        except Exception:
            return 'string'
    except Exception:
        return 'string'

def apply_format_to_range(ws, range_item, format_type, format_details):
    """Apply formatting to a specific range"""
    # Get range type from the range_item format
    if isinstance(range_item, str):
        if ':' in range_item:
            start, end = range_item.split(':')
            if re.match(r'^[A-Z]+\d+$', start) and re.match(r'^[A-Z]+\d+$', end):
                range_type = 'cells'
            elif re.match(r'^[A-Z]+$', start) and re.match(r'^[A-Z]+$', end):
                range_type = 'columns'
            elif re.match(r'^\d+$', start) and re.match(r'^\d+$', end):
                range_type = 'rows'
            else:
                raise ValueError(f"Invalid range format: {range_item}. Expected format: A:D, 1:4, or A1:D4")
        else:
            if re.match(r'^[A-Z]+\d+$', range_item):
                range_type = 'cells'
            elif re.match(r'^[A-Z]+$', range_item):
                range_type = 'columns'
            elif re.match(r'^\d+$', range_item):
                range_type = 'rows'
            else:
                raise ValueError(f"Invalid format: {range_item}. Expected format: A, 1, or A1")
    else:
        range_type = 'rows'  # If it's a number, it's a row
    
    # Convert range_item to actual Excel range based on type
    excel_range = range_item
    if range_type == 'rows':
        if ':' in str(range_item):
            start, end = map(str, range_item.split(':'))
            excel_range = f'A{start}:XFD{end}'  # XFD is the last possible Excel column
        else:
            excel_range = f'A{range_item}:XFD{range_item}'
    elif range_type == 'columns':
        if ':' in str(range_item):
            start, end = range_item.split(':')
            excel_range = f'{start}1:{end}{ws.max_row}'
        else:
            excel_range = f'{range_item}1:{range_item}{ws.max_row}'
    # For cells, range_item is already in correct format (e.g., 'A1:B2' or 'A1')

    # Apply formatting based on type
    if format_type == 'Font Colour':
        font_color = convert_to_argb(format_details.get('fontColor'))
        for row in ws[excel_range]:
            for cell in row:
                cell.font = Font(color=Color(rgb=font_color))

    elif format_type == 'Fill Colour':
        fill_color = convert_to_argb(format_details.get('fillColor'))
        for row in ws[excel_range]:
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, 
                                      end_color=fill_color, 
                                      fill_type='solid')

    elif format_type == 'Cell Number Format':
        number_format = format_details.get('numberFormat')
        format_code = {
            'number': '#,##0.00',
            'currency': '$#,##0.00',
            'accounting': '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
            'date': 'yyyy/mm/dd',
            'percentage': '0.00%',
            'custom': format_details.get('customFormat', 'General')
        }.get(number_format, format_details.get('customFormat', 'General'))
        
        for row in ws[excel_range]:
            for cell in row:
                cell.number_format = format_code

    elif format_type == 'Bold':
        for row in ws[excel_range]:
            for cell in row:
                cell.font = Font(bold=True)

    elif format_type == 'Column Width':
        width = format_details.get('columnWidth')
        # For column width, we only process column references
        if range_type == 'columns':
            if ':' in str(range_item):
                start, end = range_item.split(':')
                start_idx = column_index_from_string(start)
                end_idx = column_index_from_string(end)
                for i in range(start_idx, end_idx + 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
            else:
                ws.column_dimensions[range_item].width = width
        else:
            raise ValueError("Column width can only be applied to column ranges")

    elif format_type == 'Conditional Formatting':
        conditional = format_details.get('conditional', {})
        operator = conditional.get('operator')
        value = conditional.get('value')
        true_color = convert_to_argb(conditional.get('trueFormat', ''))
        false_color = convert_to_argb(conditional.get('falseFormat', ''))

        # Apply conditional formatting to each cell in the range
        for row in ws[excel_range]:
            for cell in row:
                cell_value = cell.value
                condition_met = False

                # Handle text-based conditions
                if operator in ['begins with', 'does not begin with', 'ends with', 'does not end with', 'contains', 'does not contain']:
                    cell_value = str(cell_value).lower()
                    value = str(value).lower()
                    
                    if operator == 'begins with':
                        condition_met = cell_value.startswith(value)
                    elif operator == 'does not begin with':
                        condition_met = not cell_value.startswith(value)
                    elif operator == 'ends with':
                        condition_met = cell_value.endswith(value)
                    elif operator == 'does not end with':
                        condition_met = not cell_value.endswith(value)
                    elif operator == 'contains':
                        condition_met = value in cell_value
                    elif operator == 'does not contain':
                        condition_met = value not in cell_value

                # Handle numeric conditions
                else:
                    try:
                        cell_value = float(cell_value) if cell_value is not None else 0
                        value = float(value)
                        
                        if operator == 'equals':
                            condition_met = cell_value == value
                        elif operator == 'does not equal':
                            condition_met = cell_value != value
                        elif operator == 'greater than':
                            condition_met = cell_value > value
                        elif operator == 'less than':
                            condition_met = cell_value < value
                        elif operator == 'greater than or equal to':
                            condition_met = cell_value >= value
                        elif operator == 'less than or equal to':
                            condition_met = cell_value <= value
                    except (ValueError, TypeError):
                        continue

                # Apply the appropriate format based on the condition
                if condition_met:
                    cell.fill = PatternFill(start_color=true_color, 
                                          end_color=true_color, 
                                          fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=false_color, 
                                          end_color=false_color, 
                                          fill_type='solid')

def format_excel_file(email, file_name, sheet_name, formatting_configs, new_file_name=''):
    """Core logic for formatting Excel files"""
    try:
        bucket = get_storage_bucket()
        
        # Try both locations for the input file
        blob = bucket.blob(f"{email}/uploaded_files/{file_name}")
        if not blob.exists():
            blob = bucket.blob(f"{email}/processed_files/{file_name}")
            if not blob.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_name}"
                }

        file_content = blob.download_as_bytes()
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active if not sheet_name else wb[sheet_name]

        # Apply each formatting configuration
        for formatting_config in formatting_configs:
            format_type = formatting_config.get('type')
            location = formatting_config.get('location', {})
            format_details = formatting_config.get('format', {})
            ranges, range_type = parse_range(location.get('range', ''))

            # Apply formatting to each range
            for range_item in ranges:
                apply_format_to_range(ws, range_item, format_type, format_details)

        # Determine output filename
        if not new_file_name:
            base_name = os.path.splitext(file_name)[0]
            output_file_name = f"formatted_{base_name}.xlsx"
        else:
            if not os.path.splitext(new_file_name)[1]:
                output_file_name = f"{new_file_name}.xlsx"
            else:
                output_file_name = new_file_name

        # Save to Firebase
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        processed_blob = bucket.blob(f"{email}/processed_files/{output_file_name}")
        processed_blob.upload_from_file(
            output_buffer, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Create DataFrame for metadata and preview
        df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name) if sheet_name else pd.read_excel(BytesIO(file_content))
        
        # Generate metadata
        metadata = {
            "sheets": {
                sheet_name or "Sheet1": {
                    "columns": df.columns.tolist(),
                    "columnTypes": {
                        col: get_column_type(df[col]) for col in df.columns
                    }
                }
            }
        }

        # Save metadata
        metadata_blob_path = f"{email}/metadata/{output_file_name}.json"
        metadata_blob = bucket.blob(metadata_blob_path)
        metadata_blob.upload_from_string(
            json.dumps(metadata),
            content_type='application/json'
        )

        # Save preview
        preview_buffer = BytesIO()
        df.head(50).to_csv(preview_buffer, index=False)
        preview_buffer.seek(0)
        preview_blob_path = f"{email}/previews/{output_file_name}/{sheet_name or 'Sheet1'}_preview.csv"
        preview_blob = bucket.blob(preview_blob_path)
        preview_blob.upload_from_file(preview_buffer, content_type='text/csv')

        return {
            "success": True,
            "message": "Formatting applied successfully.",
            "downloadUrl": processed_blob.generate_signed_url(expiration=3600, version='v4'),
            "fileName": output_file_name
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

@formatting_bp.route('/apply/', methods=['POST'])
def apply_formatting():
    """Endpoint for applying formatting to Excel files"""
    try:
        data = request.json
        email = request.headers.get('X-User-Email')
        if not email:
            return jsonify({"error": "Email is required"}), 400

        file_name = data.get('fileName')
        sheet_name = data.get('sheet')
        formatting_configs = data.get('formattingConfigs', [])
        if isinstance(formatting_configs, str):
            formatting_configs = [{"type": formatting_configs}]
        elif not isinstance(formatting_configs, list):
            formatting_configs = []

        new_file_name = data.get('newFileName', '').strip()

        result = format_excel_file(
            email=email,
            file_name=file_name,
            sheet_name=sheet_name,
            formatting_configs=formatting_configs,
            new_file_name=new_file_name
        )

        if not result.get('success'):
            return jsonify({"error": result.get('error')}), 400

        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@formatting_bp.route('/process/format/', methods=['POST'])
def process_formatting():
    """Endpoint for applying formatting to DataFrames within a process."""
    try:
        data = request.json
        if not isinstance(data, dict):
            return jsonify({"error": "Invalid input format. Expected a JSON object."}), 400

        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Validate required parameters
        required_fields = ['processId', 'tableName', 'outputTableName', 'formattingConfigs']
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        process_id = data.get('processId')
        table_name = data.get('tableName')
        output_table_name = data.get('outputTableName', '').strip()
        formatting_configs = data.get('formattingConfigs', [])

        # Get user and verify ownership
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get process and verify ownership
        process = UserProcess.query.filter_by(id=process_id, user_id=user.id).first()
        if not process:
            return jsonify({"error": "Process not found or access denied"}), 404

        # Get source DataFrame record
        source_df = DataFrame.query.filter_by(
            process_id=process_id,
            name=table_name
        ).first()
        if not source_df:
            return jsonify({"error": f"Table '{table_name}' not found in process"}), 404

        # Check if output table name already exists
        existing_df = DataFrame.query.filter_by(
            process_id=process_id,
            name=output_table_name
        ).first()

        try:
            # Generate storage location path
            storage_location = f"{email}/process/{process_id}/formatting/{output_table_name}"
            
            # Generate message for formatting operation
            message = f"Apply the specified formatting options to table {table_name}"
            
            # Create formatting step with request data and storage location
            formatting_step = FormattingStep.create_from_request(
                process_id=process_id,
                source_df_id=source_df.id,
                request_data={
                    **data,
                    'storageLocation': storage_location,
                    'message': message
                }
            )
            db.session.add(formatting_step)
            db.session.commit()

            # Process the formatting operation
            result = process_dataframe_formatting(
                email=email,
                process_id=process_id,
                source_df=source_df,
                formatting_configs=formatting_configs,
                output_table_name=output_table_name,
                existing_df=existing_df
            )

            if not result.get('success'):
                return jsonify({"error": result.get('error')}), 400

            # Add formatting step details to the response
            result.update({
                'formattingStepId': formatting_step.id,
                'processId': process_id,
                'storageLocation': storage_location,
                'sourceTable': {
                    'id': source_df.id,
                    'name': source_df.name
                },
                'description': message
            })

            return jsonify(result)

        except Exception as e:
            return jsonify({
                "error": f"An unexpected error occurred: {str(e)}",
                "formattingStepId": formatting_step.id if 'formatting_step' in locals() else None
            }), 500

    except Exception as e:
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500

def process_dataframe_formatting(email, process_id, source_df, formatting_configs, output_table_name, existing_df=None):
    """Process formatting operations on DataFrame data."""
    try:
        bucket = get_storage_bucket()

        # Read source DataFrame
        content = bucket.blob(source_df.storage_path).download_as_string()
        df = pd.read_csv(BytesIO(content))
        
        # Track formatting changes for metadata
        applied_formats = []

        # Create Excel writer
        excel_buffer = BytesIO()

        # Handle NaN/Inf values in the DataFrame first
        df = df.replace([np.inf, -np.inf], np.nan)  # Replace inf with NaN
        
        # Create a copy of the DataFrame for Excel operations
        excel_df = df.copy()
        
        # Replace NaN with empty string for Excel operations
        excel_df = excel_df.fillna('')

        with pd.ExcelWriter(
            excel_buffer, 
            engine='xlsxwriter'
        ) as writer:
            # Write DataFrame to Excel
            excel_df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Get xlsxwriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Set the nan_inf_to_errors option on the workbook
            workbook.nan_inf_to_errors = True

            # Process each formatting configuration
            for config in formatting_configs:
                format_type = config.get('type')
                location = config.get('location', {})
                format_details = config.get('format', {})
                range_str = location.get('range', '')

                # Parse the range string to get ranges and type
                ranges, range_type = parse_range(range_str)
                
                # Expand column ranges like 'A:D' to 'A2:D{last_row}' for Excel formatting
                expanded_ranges = []
                if range_type == 'columns' and len(ranges) > 1:
                    # Find first and last column
                    start_col = ranges[0]
                    end_col = ranges[-1]
                    last_row = len(df)
                    expanded_ranges.append(f"{start_col}2:{end_col}{last_row}")
                else:
                    expanded_ranges = ranges

                format_info = {
                    "type": format_type,
                    "location": location,
                    "details": format_details,
                    "rangeType": range_type,
                    "ranges": expanded_ranges
                }

                # Apply formatting based on type
                if format_type == 'Cell Number Format':
                    number_format = format_details.get('numberFormat')
                    num_format = workbook.add_format()
                    if number_format == 'number':
                        num_format.set_num_format('#,##0.00')
                        # Convert to numeric in pandas, handling NaN values
                        for range_item in ranges:
                            if range_type == 'columns':
                                col_name = df.columns[column_index_from_string(range_item) - 1]
                                df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
                                excel_df[col_name] = df[col_name].fillna('')
                    elif number_format == 'percentage':
                        num_format.set_num_format('0.00%')
                        # Convert to percentage in pandas, handling NaN values
                        for range_item in ranges:
                            if range_type == 'columns':
                                col_name = df.columns[column_index_from_string(range_item) - 1]
                                df[col_name] = pd.to_numeric(df[col_name], errors='coerce') / 100
                                excel_df[col_name] = df[col_name].fillna('')
                    elif number_format == 'date':
                        num_format.set_num_format('yyyy/mm/dd')
                        # Convert to datetime in pandas, handling NaN values
                        for range_item in ranges:
                            if range_type == 'columns':
                                col_name = df.columns[column_index_from_string(range_item) - 1]
                                df[col_name] = pd.to_datetime(df[col_name], errors='coerce')
                                excel_df[col_name] = df[col_name].fillna('')

                    # Apply Excel format
                    for range_item in ranges:
                        if range_type == 'columns':
                            col_idx = column_index_from_string(range_item) - 1
                            worksheet.set_column(col_idx, col_idx, None, num_format)
                        elif range_type == 'rows':
                            if ':' in str(range_item):
                                start, end = map(int, str(range_item).split(':'))
                                for row in range(start, end + 1):
                                    for col in range(len(df.columns)):
                                        worksheet.write(row, col, excel_df.iloc[row-1, col], num_format)
                            else:
                                row = int(range_item)
                                for col in range(len(df.columns)):
                                    worksheet.write(row, col, excel_df.iloc[row-1, col], num_format)
                        elif range_type == 'cells':
                            worksheet.conditional_format(range_item, {'type': 'cell',
                                                                   'criteria': '>=',
                                                                   'value': 0,
                                                                   'format': num_format})

                elif format_type in ['Font Colour', 'Fill Colour', 'Bold']:
                    for range_item in ranges:
                        cell_format = workbook.add_format()
                        
                        if format_type == 'Font Colour':
                            cell_format.set_font_color(format_details.get('fontColor', '').lstrip('#'))
                        elif format_type == 'Fill Colour':
                            cell_format.set_bg_color(format_details.get('fillColor', '').lstrip('#'))
                        elif format_type == 'Bold':
                            cell_format.set_bold()
                        
                        if range_type == 'columns':
                            col_idx = column_index_from_string(range_item) - 1
                            worksheet.set_column(col_idx, col_idx, None, cell_format)
                        elif range_type == 'rows':
                            if ':' in str(range_item):
                                start, end = map(int, str(range_item).split(':'))
                                for row in range(start, end + 1):
                                    for col in range(len(df.columns)):
                                        worksheet.write(row, col, excel_df.iloc[row-1, col], cell_format)
                            else:
                                row = int(range_item)
                                for col in range(len(df.columns)):
                                    worksheet.write(row, col, excel_df.iloc[row-1, col], cell_format)
                        elif range_type == 'cells':
                            if ':' in range_item:
                                start, end = range_item.split(':')
                                start_col = column_index_from_string(re.match(r'[A-Z]+', start).group()) - 1
                                start_row = int(re.search(r'\d+', start).group()) - 1
                                end_col = column_index_from_string(re.match(r'[A-Z]+', end).group()) - 1
                                end_row = int(re.search(r'\d+', end).group()) - 1
                                
                                for row in range(start_row, end_row + 1):
                                    for col in range(start_col, end_col + 1):
                                        worksheet.write(row + 1, col, excel_df.iloc[row, col], cell_format)
                            else:
                                col = column_index_from_string(re.match(r'[A-Z]+', range_item).group()) - 1
                                row = int(re.search(r'\d+', range_item).group()) - 1
                                worksheet.write(row + 1, col, excel_df.iloc[row, col], cell_format)

                elif format_type == 'Column Width':
                    width = format_details.get('columnWidth')
                    if range_type != 'columns':
                        raise ValueError("Column width can only be applied to column ranges")
                    
                    for range_item in ranges:
                        # range_item can be 'A' or 'A:C'
                        worksheet.set_column(range_item, width)

                elif format_type == 'Conditional Formatting':
                    conditional = format_details.get('conditional', {})
                    operator = conditional.get('operator')
                    value = conditional.get('value')
                    true_color = conditional.get('trueFormat', '').lstrip('#')
                    false_color = conditional.get('falseFormat', '').lstrip('#')

                    for range_item in ranges:
                        if range_type == 'columns':
                            col_idx = column_index_from_string(range_item) - 1
                            col_name = df.columns[col_idx]
                            
                            # Apply conditional formatting based on pandas operations
                            if operator in ['begins with', 'does not begin with', 'ends with', 'does not end with', 'contains', 'does not contain']:
                                # Convert column to string for text operations
                                df[col_name] = df[col_name].fillna('').astype(str)
                                
                                # Create mask based on condition
                                if operator == 'begins with':
                                    mask = df[col_name].str.lower().str.startswith(str(value).lower())
                                elif operator == 'does not begin with':
                                    mask = ~df[col_name].str.lower().str.startswith(str(value).lower())
                                elif operator == 'ends with':
                                    mask = df[col_name].str.lower().str.endswith(str(value).lower())
                                elif operator == 'does not end with':
                                    mask = ~df[col_name].str.lower().str.endswith(str(value).lower())
                                elif operator == 'contains':
                                    mask = df[col_name].str.lower().str.contains(str(value).lower(), na=False)
                                elif operator == 'does not contain':
                                    mask = ~df[col_name].str.lower().str.contains(str(value).lower(), na=False)
                            else:
                                # Handle numeric conditions
                                try:
                                    numeric_value = float(value)
                                    df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
                                    
                                    # Create mask based on numeric condition
                                    if operator == 'equals':
                                        mask = df[col_name] == numeric_value
                                    elif operator == 'does not equal':
                                        mask = df[col_name] != numeric_value
                                    elif operator == 'greater than':
                                        mask = df[col_name] > numeric_value
                                    elif operator == 'less than':
                                        mask = df[col_name] < numeric_value
                                    elif operator == 'greater than or equal to':
                                        mask = df[col_name] >= numeric_value
                                    elif operator == 'less than or equal to':
                                        mask = df[col_name] <= numeric_value
                                except (ValueError, TypeError):
                                    continue

                            # Apply conditional formatting to Excel using the mask
                            for row_idx, condition in enumerate(mask, start=1):
                                cell_format = workbook.add_format({
                                    'bg_color': true_color if condition else false_color
                                })
                                worksheet.write(row_idx, col_idx, df.iloc[row_idx-1][col_name], cell_format)
                        elif range_type == 'rows':
                            if ':' in str(range_item):
                                start, end = map(int, str(range_item).split(':'))
                                for row in range(start, end + 1):
                                    for col in range(len(df.columns)):
                                        cell_value = excel_df.iloc[row-1, col]
                                        condition_met = False
                                        
                                        # Apply the same conditional logic as for columns
                                        if operator in ['begins with', 'does not begin with', 'ends with', 'does not end with', 'contains', 'does not contain']:
                                            cell_value = str(cell_value).lower()
                                            value = str(value).lower()
                                            
                                            if operator == 'begins with':
                                                condition_met = cell_value.startswith(value)
                                            elif operator == 'does not begin with':
                                                condition_met = not cell_value.startswith(value)
                                            elif operator == 'ends with':
                                                condition_met = cell_value.endswith(value)
                                            elif operator == 'does not end with':
                                                condition_met = not cell_value.endswith(value)
                                            elif operator == 'contains':
                                                condition_met = value in cell_value
                                            elif operator == 'does not contain':
                                                condition_met = value not in cell_value
                                        else:
                                            try:
                                                cell_value = float(cell_value) if cell_value is not None else 0
                                                value = float(value)
                                                
                                                if operator == 'equals':
                                                    condition_met = cell_value == value
                                                elif operator == 'does not equal':
                                                    condition_met = cell_value != value
                                                elif operator == 'greater than':
                                                    condition_met = cell_value > value
                                                elif operator == 'less than':
                                                    condition_met = cell_value < value
                                                elif operator == 'greater than or equal to':
                                                    condition_met = cell_value >= value
                                                elif operator == 'less than or equal to':
                                                    condition_met = cell_value <= value
                                            except (ValueError, TypeError):
                                                continue
                                        
                                        cell_format = workbook.add_format({
                                            'bg_color': true_color if condition_met else false_color
                                        })
                                        worksheet.write(row, col, cell_value, cell_format)
                            else:
                                row = int(range_item)
                                for col in range(len(df.columns)):
                                    cell_value = excel_df.iloc[row-1, col]
                                    condition_met = False
                                    
                                    # Apply the same conditional logic as above
                                    if operator in ['begins with', 'does not begin with', 'ends with', 'does not end with', 'contains', 'does not contain']:
                                        cell_value = str(cell_value).lower()
                                        value = str(value).lower()
                                        
                                        if operator == 'begins with':
                                            condition_met = cell_value.startswith(value)
                                        elif operator == 'does not begin with':
                                            condition_met = not cell_value.startswith(value)
                                        elif operator == 'ends with':
                                            condition_met = cell_value.endswith(value)
                                        elif operator == 'does not end with':
                                            condition_met = not cell_value.endswith(value)
                                        elif operator == 'contains':
                                            condition_met = value in cell_value
                                        elif operator == 'does not contain':
                                            condition_met = value not in cell_value
                                    else:
                                        try:
                                            cell_value = float(cell_value) if cell_value is not None else 0
                                            value = float(value)
                                            
                                            if operator == 'equals':
                                                condition_met = cell_value == value
                                            elif operator == 'does not equal':
                                                condition_met = cell_value != value
                                            elif operator == 'greater than':
                                                condition_met = cell_value > value
                                            elif operator == 'less than':
                                                condition_met = cell_value < value
                                            elif operator == 'greater than or equal to':
                                                condition_met = cell_value >= value
                                            elif operator == 'less than or equal to':
                                                condition_met = cell_value <= value
                                        except (ValueError, TypeError):
                                            continue
                                    
                                    cell_format = workbook.add_format({
                                        'bg_color': true_color if condition_met else false_color
                                    })
                                    worksheet.write(row, col, cell_value, cell_format)
                        elif range_type == 'cells':
                            # For cell ranges, apply conditional formatting directly
                            worksheet.conditional_format(range_item, {
                                'type': 'cell',
                                'criteria': operator,
                                'value': value,
                                'format': workbook.add_format({'bg_color': true_color}),
                                'multi_range': range_item
                            })

                applied_formats.append(format_info)

        # Generate storage paths
        csv_storage_path = f"{email}/process/{process_id}/dataframes/{output_table_name}.csv"
        excel_storage_path = f"{email}/process/{process_id}/formatted/{output_table_name}.xlsx"

        try:
            # Initialize blob variables
            csv_blob = None
            excel_blob = None
            metadata_blob = None
            preview_blob = None
            
            # Save CSV version for process storage (using original DataFrame with NaN values)
            csv_buffer = BytesIO()
            df.to_csv(csv_buffer, index=False)
            csv_buffer.seek(0)
            csv_blob = bucket.blob(csv_storage_path)
            csv_blob.upload_from_file(csv_buffer, content_type='text/csv')

            # Save formatted Excel version (using DataFrame with empty strings)
            excel_buffer.seek(0)
            excel_blob = bucket.blob(excel_storage_path)
            excel_blob.upload_from_file(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Create or update DataFrame record
            if existing_df:
                existing_df.row_count = len(df)
                existing_df.column_count = len(df.columns)
                existing_df.updated_at = datetime.now(timezone.utc)
                existing_df.storage_path = csv_storage_path
                dataframe_record = existing_df
            else:
                dataframe_record = DataFrame.create_from_pandas(
                    df=df,
                    process_id=process_id,
                    name=output_table_name,
                    email=email,
                    storage_path=csv_storage_path,
                    user_id=source_df.user_id
                )
                db.session.add(dataframe_record)

            # Store the formatted Excel path in metadata
            metadata = {
                "type": "formatted_table",
                "description": f"Formatted table from {source_df.name}",
                "processId": process_id,
                "createdAt": (existing_df.created_at if existing_df else datetime.now(timezone.utc)).strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                "updatedAt": datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                "rowCount": len(df),
                "columnCount": len(df.columns),
                "columns": df.columns.tolist(),
                "columnTypes": {col: str(df[col].dtype) for col in df.columns},
                "formatting": {
                    "appliedFormats": applied_formats,
                    "sourceTable": {
                        "id": source_df.id,
                        "name": source_df.name
                    },
                    "formattedExcelPath": excel_storage_path
                }
            }

            metadata_path = f"{email}/process/{process_id}/metadata/{output_table_name}.json"
            metadata_blob = bucket.blob(metadata_path)
            metadata_blob.upload_from_string(
                json.dumps(metadata, indent=2),
                content_type='application/json'
            )

            # Save preview data (using original DataFrame with NaN values)
            preview_df = df.head(50)
            preview_buffer = BytesIO()
            preview_df.to_csv(preview_buffer, index=False)
            preview_buffer.seek(0)
            
            preview_path = f"{email}/process/{process_id}/previews/{output_table_name}_preview.csv"
            preview_blob = bucket.blob(preview_path)
            preview_blob.upload_from_file(preview_buffer, content_type='text/csv')

            # Commit the database changes
            db.session.commit()

            # Add download URL for formatted Excel file
            download_url = excel_blob.generate_signed_url(expiration=3600, version='v4')

            return {
                "success": True,
                "message": f"Table formatted successfully as '{output_table_name}'",
                "id": dataframe_record.id,
                "name": output_table_name,
                "rowCount": len(df),
                "columnCount": len(df.columns),
                "isUpdate": existing_df is not None,
                "downloadUrl": download_url
            }

        except Exception as e:
            db.session.rollback()
            # Clean up uploaded files if they exist
            for blob in [csv_blob, excel_blob, metadata_blob, preview_blob]:
                if blob and blob.exists():
                    blob.delete()
            raise e

    except Exception as e:
        raise Exception(f"Error formatting table: {str(e)}")

@formatting_bp.route('/process/include/', methods=['POST'])
def include_formatting_step():
    """Endpoint to include a formatting step in process execution."""
    try:
        data = request.json
        if not isinstance(data, dict):
            return jsonify({"error": "Invalid input format. Expected a JSON object."}), 400

        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Validate required parameters
        required_fields = ['formattingStepId', 'processId']
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        formatting_step_id = data.get('formattingStepId')
        process_id = data.get('processId')

        # Get user and verify ownership
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get process and verify ownership
        process = UserProcess.query.filter_by(id=process_id, user_id=user.id).first()
        if not process:
            return jsonify({"error": "Process not found or access denied"}), 404

        # Get formatting step and verify it belongs to the process
        formatting_step = FormattingStep.query.filter_by(
            id=formatting_step_id,
            process_id=process_id
        ).first()
        if not formatting_step:
            return jsonify({"error": "Formatting step not found or does not belong to the process"}), 404

        # Update the include_in_process field
        formatting_step.include_in_process = True
        formatting_step.updated_at = datetime.now(timezone.utc)
        
        try:
            db.session.add(formatting_step)
            db.session.commit()

            return jsonify({
                "success": True,
                "message": "Formatting step included in process execution",
                "formattingStep": formatting_step.to_dict()
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": f"Failed to update formatting step: {str(e)}"
            }), 500

    except Exception as e:
        return jsonify({
            "error": f"An unexpected error occurred: {str(e)}"
        }), 500

@formatting_bp.route('/process/step/<formatting_step_id>', methods=['DELETE'])
def delete_formatting_step(formatting_step_id):
    """Endpoint to delete a formatting step."""
    try:
        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Get user
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get formatting step and verify ownership through process
        formatting_step = FormattingStep.query.join(
            UserProcess, FormattingStep.process_id == UserProcess.id
        ).filter(
            FormattingStep.id == formatting_step_id,
            UserProcess.user_id == user.id
        ).first()

        if not formatting_step:
            return jsonify({"error": "Formatting step not found or access denied"}), 404

        try:
            # Get storage location from formatting step
            storage_location = formatting_step.storage_location
            
            # Delete files from Firebase if storage location exists
            if storage_location:
                bucket = get_storage_bucket()
                # List all blobs with the storage location prefix
                blobs = bucket.list_blobs(prefix=storage_location)
                for blob in blobs:
                    blob.delete()

            # Delete the formatting step from database
            db.session.delete(formatting_step)
            db.session.commit()

            return jsonify({
                "success": True,
                "message": "Formatting step deleted successfully",
                "deletedId": formatting_step_id
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": f"Failed to delete formatting step: {str(e)}"
            }), 500

    except Exception as e:
        return jsonify({
            "error": f"An unexpected error occurred: {str(e)}"
        }), 500

@formatting_bp.route('/process/<process_id>/steps', methods=['GET'])
def get_process_formatting_steps(process_id):
    """Get all formatting steps for a process."""
    try:
        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Get user
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get process and verify ownership
        process = UserProcess.query.filter_by(id=process_id, user_id=user.id).first()
        if not process:
            return jsonify({"error": "Process not found or access denied"}), 404

        # Get all formatting steps for the process
        formatting_steps = FormattingStep.query.filter_by(process_id=process_id, include_in_process=True).all()
        
        # Convert to dictionary format
        steps_data = []
        for step in formatting_steps:
            step_dict = step.to_dict()
            step_dict['title'] = "Formatting"  # Add title field
            # Add source DataFrame details
            if step.source_dataframe:
                step_dict['sourceDataframe'] = {
                    'id': step.source_dataframe.id,
                    'name': step.source_dataframe.name,
                    'rowCount': step.source_dataframe.row_count,
                    'columnCount': step.source_dataframe.column_count
                }
            steps_data.append(step_dict)

        return jsonify({
            "success": True,
            "process": {
                "id": process.id,
                "name": process.process_name
            },
            "formattingSteps": steps_data,
            "summary": {
                "totalSteps": len(steps_data),
                "includedInProcess": sum(1 for step in steps_data if step.get('includeInProcess', False))
            }
        })

    except Exception as e:
        return jsonify({
            "error": f"An unexpected error occurred: {str(e)}"
        }), 500

@formatting_bp.route('/process/step/<formatting_step_id>', methods=['PUT'])
def update_formatting_step(formatting_step_id):
    """Endpoint to update a formatting step configuration."""
    try:
        data = request.json
        if not isinstance(data, dict):
            return jsonify({"error": "Invalid input format. Expected a JSON object."}), 400

        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Validate required parameters
        if 'configuration' not in data:
            return jsonify({"error": "Missing required field: configuration"}), 400

        # Get user
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get formatting step and verify ownership through process
        formatting_step = FormattingStep.query.join(
            UserProcess, FormattingStep.process_id == UserProcess.id
        ).filter(
            FormattingStep.id == formatting_step_id,
            UserProcess.user_id == user.id
        ).first()

        if not formatting_step:
            return jsonify({"error": "Formatting step not found or access denied"}), 404

        try:
            config = data['configuration']

            if "formattingConfigs" in config:
                row_count = None
                if formatting_step.source_dataframe:
                    row_count = formatting_step.source_dataframe.row_count  # e.g., 2178

                for fc in config["formattingConfigs"]:
                    if "location" in fc and "range" in fc["location"]:
                        original_range = fc["location"]["range"]
                        # If the range is column-only (like "F:H"), expand with rows
                        if ":" in original_range and not any(ch.isdigit() for ch in original_range):
                            start_col, end_col = original_range.split(":")
                            end_row = row_count + 1 if row_count else 1048576  # include header offset
                            fc["location"]["range"] = f"{start_col}2:{end_col}{end_row}"


            # Update configuration
            formatting_step.configuration = config
            formatting_step.updated_at = datetime.now(timezone.utc)
            
            db.session.add(formatting_step)
            db.session.commit()

            return jsonify({
                "success": True,
                "message": "Formatting step configuration updated successfully",
                "formattingStep": formatting_step.to_dict()
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": f"Failed to update formatting step: {str(e)}"
            }), 500

    except Exception as e:
        return jsonify({
            "error": f"An unexpected error occurred: {str(e)}"
        }), 500

@formatting_bp.route('/process/step/<formatting_step_id>', methods=['GET'])
def get_formatting_step(formatting_step_id):
    """
    Get formatting step details by ID.
    
    Headers:
        X-User-Email: User's email for authentication
    
    Parameters:
        formatting_step_id: ID of the formatting step to retrieve
    """
    try:
        email = request.headers.get("X-User-Email")
        if not email:
            return jsonify({"error": "Email is required in the headers."}), 400

        # Get user
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Get formatting step and verify ownership through process
        formatting_step = FormattingStep.query.join(
            UserProcess, FormattingStep.process_id == UserProcess.id
        ).filter(
            FormattingStep.id == formatting_step_id,
            UserProcess.user_id == user.id
        ).first()

        if not formatting_step:
            return jsonify({"error": "Formatting step not found or access denied"}), 404

        # Get process details
        process = UserProcess.query.get(formatting_step.process_id)

        # Build response with detailed information
        response_data = {
            "success": True,
            "formattingStep": {
                **formatting_step.to_dict(),  # Base formatting step info
                "sourceDataframe": None,  # Will be populated if source exists
                "process": {
                    "id": process.id,
                    "name": process.process_name,
                    "createdAt": process.created_at.isoformat() if process.created_at else None,
                    "updatedAt": process.updated_at.isoformat() if process.updated_at else None
                }
            }
        }

        # Add source DataFrame details if available
        if formatting_step.source_dataframe:
            source_df = formatting_step.source_dataframe
            response_data["formattingStep"]["sourceDataframe"] = {
                "id": source_df.id,
                "name": source_df.name,
                "rowCount": source_df.row_count,
                "columnCount": source_df.column_count,
                "createdAt": source_df.created_at.isoformat() if source_df.created_at else None,
                "updatedAt": source_df.updated_at.isoformat() if source_df.updated_at else None
            }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({
            "error": f"An unexpected error occurred: {str(e)}"
        }), 500


